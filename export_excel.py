"""Create hasil_screening.xlsx with decisions, audit metadata, and charts."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def export(input_path: Path, output_path: Path) -> None:
    rows = load_jsonl(input_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Screening decisions"
    headers = ["Record ID", "Title", "Decision", "Confidence", "Rationale", "Criteria met", "Criteria failed", "Source", "Latency ms", "Router trace IDs", "Criteria hash"]
    sheet.append(headers)
    for item in rows:
        sheet.append([item["record_id"], item["title"], item["decision"], item["confidence"], item["rationale"],
                      "; ".join(item.get("criteria_met", [])), "; ".join(item.get("criteria_failed", [])), item["source"],
                      item.get("latency_ms", 0), "; ".join(item.get("router_trace_ids", [])), item["criteria_hash"]])
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="5F55C7")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [22, 58, 14, 12, 64, 28, 28, 24, 12, 35, 22]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    if rows:
        table = Table(displayName="ScreeningLog", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        sheet.add_table(table)

    summary = workbook.create_sheet("Summary")
    summary.append(["Decision", "Count"])
    for decision in ("include", "exclude", "review"):
        summary.append([decision.title(), sum(row["decision"] == decision for row in rows)])
    summary.append([])
    summary.append(["Metric", "Value"])
    summary.append(["Total records", len(rows)])
    summary.append(["Mean confidence", sum(float(row["confidence"]) for row in rows) / max(len(rows), 1)])
    summary.append(["Router decisions", sum(row["source"] == "9router-ensemble" for row in rows)])
    pie = PieChart()
    pie.title = "Screening disposition"
    pie.add_data(Reference(summary, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    pie.set_categories(Reference(summary, min_col=1, min_row=2, max_row=4))
    summary.add_chart(pie, "D2")
    bar = BarChart()
    bar.title = "Records by decision"
    bar.add_data(Reference(summary, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    bar.set_categories(Reference(summary, min_col=1, min_row=2, max_row=4))
    summary.add_chart(bar, "D18")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


if __name__ == "__main__":
    cli = argparse.ArgumentParser()
    cli.add_argument("input", help="screening.jsonl")
    cli.add_argument("--output", default="outputs/slr_run/hasil_screening.xlsx")
    args = cli.parse_args()
    export(Path(args.input), Path(args.output))
